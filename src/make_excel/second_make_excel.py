from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl import Workbook
import String
from datetime import datetime


# 두번째 요구의 sheet 이름
second_1차상담_직탐미설 = String.second_1차상담_직탐미설
second_1차상담_신입생세미나 = String.second_1차상담_신입생세미나
second_1차상담_자기주도 = String.second_1차상담_자기주도
second_1차상담_진로집단상담 = String.second_1차상담_진로집단상담
second_1차상담_단기직무체험 = String.second_1차상담_단기직무체험

second_2차상담_직탐미설 = String.second_2차상담_직탐미설
second_2차상담_포트워크숍 = String.second_2차상담_포트워크숍
second_2차상담_자기주도 = String.second_2차상담_자기주도
second_2차상담_진로집단상담 = String.second_2차상담_진로집단상담
second_2차상담_단기직무체험 = String.second_2차상담_단기직무체험

second_3차상담_포트폴리오컨설팅 = String.second_3차상담_포트폴리오컨설팅

# input excel file 의 시트이름
직업탐색과미래설계 = String.second_직업탐색과미래설계
신입생세미나 = String.second_신입생세미나
포트폴리오워크숍 = String.second_포트폴리오워크숍
포트폴리오컨설팅 = String.second_포트폴리오컨설팅
자기주도 = String.second_자기주도
진로집단상담 = String.second_진로집단상담
단기직무체험 = String.second_단기직무체험

first_상담 = String.second_1차_상담
second_상담 = String.second_2차_상담
third_상담 = String.second_3차_상담


def make_second_excel(write_excel_file, read_excel_file):
    """_summary_

    Args:
        write_excel_file (Excel):  쓰고싶은  엑셀의 시트
        read_excel_file (Excel): 읽고 싶은 엑셀 
    """
    # 직업탐색과미래설계
    make_sheet_finish(write_excel_file, read_excel_file,
                      직업탐색과미래설계, second_1차상담_직탐미설, first_상담)
    make_sheet_finish(write_excel_file, read_excel_file,
                      직업탐색과미래설계, second_2차상담_직탐미설, second_상담)

    # 신입생세미나
    make_sheet_finish(write_excel_file, read_excel_file,
                      신입생세미나, second_1차상담_신입생세미나, first_상담)

    # 포트폴리오워크숍
    make_sheet_finish(write_excel_file, read_excel_file,
                      포트폴리오워크숍, second_2차상담_포트워크숍, second_상담)

    # 자기주도
    make_sheet_finish(write_excel_file, read_excel_file,
                      자기주도, second_1차상담_자기주도, first_상담)
    make_sheet_finish(write_excel_file, read_excel_file,
                      자기주도, second_2차상담_자기주도, second_상담)

    # 진로집단상담
    make_sheet_finish(write_excel_file, read_excel_file,
                      진로집단상담, second_1차상담_진로집단상담, first_상담)
    make_sheet_finish(write_excel_file, read_excel_file,
                      진로집단상담, second_2차상담_진로집단상담, second_상담)

    # 단기직무체험
    make_sheet_finish(write_excel_file, read_excel_file,
                      단기직무체험, second_1차상담_단기직무체험, first_상담)
    make_sheet_finish(write_excel_file, read_excel_file,
                      단기직무체험, second_2차상담_단기직무체험, second_상담)

    # 포트폴리오컨설팅
    make_sheet_finish(write_excel_file, read_excel_file,
                      포트폴리오컨설팅, second_3차상담_포트폴리오컨설팅, third_상담)


def make_sheet_finish(write_excel_file, read_excel_file, input_select_sheet_name, name_select_sheet, type_consultation):
    """_summary_

    Args:
        write_excel_file (Excel):  쓰고싶은  엑셀의 시트
        read_excel_file (Excel): 읽고 싶은 엑셀 
        name_select_sheet (string):  시트의 이름
    """
    #  해당시트의  학생정보가  시작한는 행번호
    select_sheet_student_info_base_row = 3

    # read_excel_file 파일의 sheet 가져오기
    read_select_sheet = read_excel_file[input_select_sheet_name]

    # 2번째 sheet 열기
    sheet_second_write_excel = write_excel_file[name_select_sheet]

    # 해당하는 시트의 행의 길이
    read_select_sheet_max_row = read_select_sheet.max_row

    # 선택 시트의 정보가있는 2 번째 row에대한  행을 다 가져온다
    # no, 신청경로, 참여구분 ,....., 2차상담, 포트폴리오
    # 2번 row 에  해당하는 정보가있다.
    sheet_info_row_two = read_select_sheet[2]

    name = None
    contact = None
    student_id = None
    company_name = None
    first_consultation_time_full = None
    second_consultation_time_full = None
    third_consultation_time_full = None

    for cell in sheet_info_row_two:
        if cell.value == String.second_성명:
            name = cell.column_letter
        if cell.value == String.second_연락처:
            contact = cell.column_letter
        if cell.value == String.second_학번:
            student_id = cell.column_letter
        if cell.value == String.second_삼당사명:
            company_name = cell.column_letter
        if cell.value == String.second_삼당사:
            company_name = cell.column_letter
        if cell.value == String.second_1차_상담:
            first_consultation_time_full = cell.column_letter
        if cell.value == String.second_2차_상담:
            second_consultation_time_full = cell.column_letter
        if cell.value == String.second_3차_상담:
            third_consultation_time_full = cell.column_letter

    # 해당 읽고싶어하는 sheet의  모든 학생 탐색
    for select_sheet_row in range(select_sheet_student_info_base_row, read_select_sheet_max_row + 1):
        # 이름
        value_in_name = read_select_sheet[str(
            name) + str(select_sheet_row)].value

        # 연락처
        value_in_contact = read_select_sheet[str(
            contact) + str(select_sheet_row)].value

        # 학번
        value_in_id = read_select_sheet[str(
            student_id) + str(select_sheet_row)].value

        # 상담사
        value_in_company_name = read_select_sheet[str(
            company_name) + str(select_sheet_row)].value

        value_in_consultation_date, value_in_consultation_hour, value_in_consultation_minute, value_in_consultation_time = None, None, None, None

        if (type_consultation == first_상담):
            # 1차상담 2023-04-12 10:00
            value_in_first_consultation_time_full = read_select_sheet[str(
                first_consultation_time_full) + str(select_sheet_row)].value

            # 1차상담 2023-04-12 10:00 을 원하는  데이터 추출
            data_result = extract_datetime_parts(
                value_in_first_consultation_time_full)

            # 상담일자 20230412, 상담시 10, 상담분 00, 상담시간 60분
            value_in_consultation_date, value_in_consultation_hour, value_in_consultation_minute, value_in_consultation_time = data_result

        elif (type_consultation == second_상담):
            # 2차상담 2023-04-12 10:00
            value_in_first_consultation_time_full = read_select_sheet[str(
                second_consultation_time_full) + str(select_sheet_row)].value

            # 2차상담 2023-04-12 10:00 을 원하는  데이터 추출
            data_result = extract_datetime_parts(
                value_in_first_consultation_time_full)

            # 상담일자 20230412, 상담시 10, 상담분 00, 상담시간 60분
            value_in_consultation_date, value_in_consultation_hour, value_in_consultation_minute, value_in_consultation_time = data_result
        elif (type_consultation == third_상담):
            # 3차상담 2023-04-12 10:00
            value_in_first_consultation_time_full = read_select_sheet[str(
                third_consultation_time_full) + str(select_sheet_row)].value

            # 3차상담 2023-04-12 10:00 을 원하는  데이터 추출
            data_result = extract_datetime_parts(
                value_in_first_consultation_time_full)

            # 상담일자 20230412, 상담시 10, 상담분 00, 상담시간 60분
            value_in_consultation_date, value_in_consultation_hour, value_in_consultation_minute, value_in_consultation_time = data_result

        # 엑셀에 원하는 정보 쓰기
        # 이름
        sheet_second_write_excel['B' +
                                 str(select_sheet_row - 1)] = value_in_name
        # 연락처
        sheet_second_write_excel['C' +
                                 str(select_sheet_row - 1)] = value_in_contact
        # 학번
        sheet_second_write_excel['D' +
                                 str(select_sheet_row - 1)] = value_in_id
        # 상담사
        sheet_second_write_excel['J' +
                                 str(select_sheet_row - 1)] = value_in_company_name
        # 상담일자
        sheet_second_write_excel['L' +
                                 str(select_sheet_row - 1)] = value_in_consultation_date
        sheet_second_write_excel['M' +
                                 str(select_sheet_row - 1)] = value_in_consultation_hour
        sheet_second_write_excel['N' +
                                 str(select_sheet_row - 1)] = value_in_consultation_minute
        sheet_second_write_excel['O' +
                                 str(select_sheet_row - 1)] = value_in_consultation_time
        # 순번
        number = select_sheet_row - 2
        sheet_second_write_excel['A' +
                                 str(select_sheet_row - 1)] = number


def extract_datetime_parts(date_time_str):
    if (date_time_str != None):
        # 입력된 문자열을 datetime 객체로 파싱합니다.
        date_time = datetime.strptime(str(date_time_str), '%Y-%m-%d %H:%M:%S')

        # 연도, 월, 일, 시간, 분을 추출합니다.
        consultation_date = date_time.strftime('%Y%m%d')
        consultation_hour = date_time.strftime('%H')
        consultation_minute = date_time.strftime('%M')
        consultation_time = "60"

        return consultation_date, consultation_hour, consultation_minute, consultation_time
    return None, None, None, None
