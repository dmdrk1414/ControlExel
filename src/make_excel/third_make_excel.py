from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl import Workbook
import String

third_신청경로 = String.third_신청경로
third_학번 = String.third_학번
third_성명 = String.third_성명
third_연락처 = String.third_연락처
third_학년 = String.third_학년
third_성별 = String.third_성별
third_소속학과 = String.third_소속학과
third_이수구분 = String.third_이수구분
third_빌드업신청 = String.third_빌드업신청

third_build_신청일자 = String.third_build_신청일자
third_build_처리상태 = String.third_build_처리상태


def make_third_excel(write_excel_file, read_excel_file, select_write_sheet):
    """_summary_

    Args:
        write_excel_file (Excel):  쓰고싶은  엑셀의 시트
        read_excel_file (Excel): 읽고 싶은 엑셀 
    """

    # read_excel_file 파일의 sheet 가져오기
    read_select_sheet = read_excel_file[select_write_sheet]

    # read_excel_file 파일의 빌드업신청 가져오기
    read_build_up_sheet = read_excel_file[third_빌드업신청]

    # 원하는 시트을 연다.
    sheet_third_write_excel = write_excel_file[select_write_sheet]

    # 해당하는 시트의 행의 길이
    read_select_sheet_max_row = read_select_sheet.max_row

    # 선택 시트의 정보가있는 2 번째 row에대한  행을 다 가져온다
    # no, 신청경로, 참여구분 ,....., 2차상담, 포트폴리오
    # 2번 row 에  해당하는 정보가있다.
    sheet_info_row_two = read_select_sheet[2]

    #  해당시트의  학생정보가  시작한는 행번호
    select_sheet_student_info_base_row = 3

    application_route = None
    student_id = None
    name = None
    contact = None
    grade = None
    gender = None
    department = None
    completion_status = None

    for cell in sheet_info_row_two:
        if cell.value == third_신청경로:
            application_route = cell.column_letter
        if cell.value == third_학번:
            student_id = cell.column_letter
        if cell.value == third_성명:
            name = cell.column_letter
        if cell.value == third_연락처:
            contact = cell.column_letter
        if cell.value == third_학년:
            grade = cell.column_letter
        if cell.value == third_성별:
            gender = cell.column_letter
        if cell.value == third_소속학과:
            department = cell.column_letter
        if cell.value == third_이수구분:
            completion_status = cell.column_letter

    for select_sheet_row in range(select_sheet_student_info_base_row, read_select_sheet_max_row + 1):
        # 신청경로
        value_in_application_route = select_write_sheet

        # 학번
        value_in_student_id = read_select_sheet[str(
            student_id) + str(select_sheet_row)].value

        # 성명
        value_in_name = read_select_sheet[str(
            name) + str(select_sheet_row)].value

        # 연락처
        value_in_contact = read_select_sheet[str(
            contact) + str(select_sheet_row)].value

        # 학년
        value_in_grade = read_select_sheet[str(
            grade) + str(select_sheet_row)].value

        # 성별
        value_in_gender = read_select_sheet[str(
            gender) + str(select_sheet_row)].value

        # 소속학과
        value_in_department = read_select_sheet[str(
            department) + str(select_sheet_row)].value

        # 이수여부
        value_in_completion_status = read_select_sheet[str(
            completion_status) + str(select_sheet_row)].value

        # ======== 쓰기 ====================
        # 순번 쓰기
        number = select_sheet_row - 2
        sheet_third_write_excel['A' +
                                str(select_sheet_row)] = number

        # 신청경로 쓰기
        sheet_third_write_excel['B' +
                                str(select_sheet_row)] = value_in_application_route

        # 학번 쓰기
        sheet_third_write_excel['C' +
                                str(select_sheet_row)] = value_in_student_id

        # 성명 쓰기
        sheet_third_write_excel['D' +
                                str(select_sheet_row)] = value_in_name

        # 연락처 쓰기
        sheet_third_write_excel['E' +
                                str(select_sheet_row)] = value_in_contact

        # 학년 쓰기
        sheet_third_write_excel['F' +
                                str(select_sheet_row)] = value_in_grade

        # 성별 쓰기
        sheet_third_write_excel['G' +
                                str(select_sheet_row)] = value_in_gender

        # 소속학과 쓰기
        sheet_third_write_excel['H' +
                                str(select_sheet_row)] = value_in_department

        # 이수여부 쓰기
        sheet_third_write_excel['I' +
                                str(select_sheet_row)] = value_in_completion_status

        value_in_processing_status, value_in_application_data = build_up_part(
            read_build_up_sheet, value_in_student_id)

        # 빌드업신청 쓰기
        sheet_third_write_excel['J' +
                                str(select_sheet_row)] = value_in_processing_status

        # 신청일자 쓰기
        sheet_third_write_excel['L' +
                                str(select_sheet_row)] = value_in_application_data


def build_up_part(read_build_up_sheet, value_in_want_sheet_student_id):
    """ build_up sheet의 해당 학번이 있는  곳의 처리상태, 신청일자의 값을 가져온다.
    """
    sheet_info_row_two = read_build_up_sheet[2]

    # build 시트에서 학번이 있는 열
    build_student_id = None

    # 해당하는 시트의 행의 길이
    read_select_sheet_max_row = read_build_up_sheet.max_row

    # build sheet의 학번 정보
    value_in_build_student_id = None

    #  원하는 학번이있는 행의 정보
    row_want_build_same_id = None

    # 신청일자, 처리상태 의 열
    application_data = None
    processing_status = None

    # 신청일자, 처리상태 의 값
    value_in_processing_status = None
    value_in_application_data = None

    # 신청일자, 처리상태 의 열 찾기
    for cell in sheet_info_row_two:
        if cell.value == third_build_처리상태:
            processing_status = cell.column_letter
        if cell.value == third_build_신청일자:
            application_data = cell.column_letter

    #  build up 시트의  학생정보가  시작한는 행번호
    select_sheet_student_info_base_row = 3
    for cell in sheet_info_row_two:
        if cell.value == third_학번:
            build_student_id = cell.column_letter

    # 원하는시트와 build_up  시트의  학번이 동일한지 확인한후 그행을 가져온다.
    for select_sheet_row in range(select_sheet_student_info_base_row, read_select_sheet_max_row + 1):
        try:
            value_in_build_student_id = read_build_up_sheet[str(
                build_student_id) + str(select_sheet_row)].value

            if (str(value_in_build_student_id) == str(value_in_want_sheet_student_id)):
                row_want_build_same_id = select_sheet_row
                break
        except ValueError as e:
            # ValueError 예외가 발생한 경우 이 블록이 실행됩니다.
            print("ValueError 예외가 발생했습니다:", e)

    try:
        # build_up sheet의 해당 학번이 있는  곳의 처리상태, 신청일자의 값을 가져온다.
        value_in_processing_status = read_build_up_sheet[str(
            processing_status) + str(row_want_build_same_id)].value
        value_in_application_data = read_build_up_sheet[str(
            application_data) + str(row_want_build_same_id)].value
    except ValueError as e:
        value_in_processing_status = ""
        value_in_application_data = ""
        print(str(value_in_want_sheet_student_id) +
              " 이 학생은 빌드업신청 sheet에서 학번이 없습니다.")

    return value_in_processing_status, value_in_application_data
