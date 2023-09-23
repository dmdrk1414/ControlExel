from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from make_base_sheet import second_make_base  # 모듈을 현재 디렉토리에서 import합니다.
from make_base_sheet import third_make_base  # 모듈을 현재 디렉토리에서 import합니다.
from make_excel import second_make_excel  # 모듈을 현재 디렉토리에서 import합니다.
from make_excel import third_make_excel  # 모듈을 현재 디렉토리에서 import합니다.


# 현재 날짜와 시간 가져오기
today = datetime.now()
# 월과 일 얻기
current_month = today.month  # 현재 월
current_day = today.day      # 현재 일
# 재맞고 빌드업 참여현황(9.23기준)
inputFile = "재맞고 빌드업 참여현황(" + str(current_month) + "." + \
    str(current_day) + "기준)" + ".xlsx"
outputFile = "output_전산망개발" + ".xlsx"

input_xcel_src = "../input_excel_file/" + inputFile
output_xcel_src = "../output_excel_file/" + outputFile

# 열의 넓이를 적어라 2번째 요구
numWidthSecondColum = 22

# 엑셀파일 쓰기
write_wb = Workbook()

# read xcel file
# data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook(
    input_xcel_src, data_only=True)


def makeSheet(str):
    # 이름이 있는 시트를 생성
    write_wb.create_sheet(str)


def callMakeExcel():
    # 전산망 시트 읽기 열기
    build_up_networkSheet = load_wb['빌드업신청']

    # 2번째 sheet 쓰기 열기
    write_ws = write_wb['2번째']

    # 2번째 sheet에 요구사항을 실행한다
    second_make_excel.makeSecondExcel(build_up_networkSheet, write_ws, load_wb)

    # # 3번째 sheet 쓰기 열기
    # write_ws = write_wb['3번째']

    # # 3번째 sheet에 요구사항을 실행한다
    # third_make_excel.makeThirdExcel(computerNetworkSheet, write_ws, load_wb)


# 요구하는 시트의 베이스 엑셀을 만든다.
def callMakeBaseExcel():
    # 2번째 sheet 열기
    write_ws = write_wb['2번째']

    # 2번째요구의  시트의 베이스 엑셀을 만든다.
    second_make_base.makeSecondBase(write_ws, numWidthSecondColum)

    # # 3번째 sheet 열기
    # write_ws = write_wb['3번째']

    # # 3번째요구의  시트의 베이스 엑셀을 만든다.
    # third_make_base.makeThirdBase(write_ws, numWidthSecondColum)


def run():
    # 기본 시트 제거
    default_sheet = write_wb.active
    write_wb.remove(default_sheet)

    # make second sheet
    makeSheet("2번째")
    # makeSheet("3번째")

    # 요구하는 시트의 베이스 엑셀을 만든다.
    callMakeBaseExcel()

    # 요구사항을 실행한다
    callMakeExcel()

    write_wb.save(output_xcel_src)


if __name__ == "__main__":
    run()
