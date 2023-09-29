from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from src.make_base_sheet import first_make_base  # 모듈을 현재 디렉토리에서 import합니다.
from src.make_excel import first_make_excel  # 모듈을 현재 디렉토리에서 import합니다.
import String


def make_first_request_excel():
    # 현재 날짜와 시간 가져오기
    today = datetime.now()
    # 월과 일 얻기
    current_month = today.month  # 현재 월
    current_day = today.day      # 현재 일
    # 재맞고 빌드업 참여현황(9.23기준)
    inputFile = String.NAME_FIRST_INPUT_FILE+"(" + str(current_month) + "." + \
        str(current_day) + "기준)" + ".xlsx"

    outputFile = String.NAME_FIRST_OUT_FILE + ".xlsx"

    input_xcel_src = "./input_excel_file/" + inputFile
    output_xcel_src = "./output_excel_file/" + outputFile

    # 엑셀파일 쓰기
    write_excel = Workbook()

    # read xcel file
    # data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
    load_wb = load_workbook(
        input_xcel_src, data_only=True)

    # 기본 시트 제거
    default_sheet = write_excel.active
    write_excel.remove(default_sheet)

    # 1 번째 sheet 만들기
    write_excel.create_sheet('1번째')

    # 1번째 sheet 열기
    sheet_first_write_excel = write_excel['1번째']

    # 1번째요구의  시트의 베이스 엑셀을 만든다.
    first_make_base.makeSecondBase(
        sheet_first_write_excel, String.num_width_colum)

    # 빌드업신청 읽기 열기
    build_up_network_sheet = load_wb['빌드업신청']

    # 1번째 요구사항 excel의 1번째 sheet에 요구사항을 실행한다
    first_make_excel.makeSecondExcel(
        build_up_network_sheet, sheet_first_write_excel, load_wb)

    # 1 번째 요구에맞는 엑셀만들기
    write_excel.save(output_xcel_src)
