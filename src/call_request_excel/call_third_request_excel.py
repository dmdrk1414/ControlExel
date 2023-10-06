from openpyxl import load_workbook
from openpyxl import Workbook
# 모듈을 현재 디렉토리에서 import합니다.
from src.make_base_sheet.third_make_base import make_third_base
from src.make_excel.third_make_excel import make_third_excel
import String


# make 3
def make_third_request_excel():
    # second file 이름
    inputFile = String.NAME_THIRD_INPUT_FILE + ".xlsx"

    outputFile = String.NAME_THIRD_OUT_FILE + ".xlsx"

    input_xcel_src = "./input_excel_file/" + inputFile
    output_xcel_src = "./output_excel_file/" + outputFile

    # 엑셀파일 쓰기
    write_excel = Workbook()

    # read xcel file
    # data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
    load_excel_file = load_workbook(
        input_xcel_src, data_only=True)

    # 기본 시트 제거
    default_sheet = write_excel.active
    write_excel.remove(default_sheet)

    select_write_sheet = input("만들고 싶은 시트 이름을 적으세요(ex 포폴컨설팅) : ")

    # 3번째 sheet 만들기
    write_excel.create_sheet(select_write_sheet)

    # 3번째 sheet 열기
    sheet_second_write_excel = write_excel[select_write_sheet]

    # 3번째요구의  시트의 베이스 엑셀을 만든다.
    make_third_base(
        sheet_second_write_excel, String.num_width_colum, select_write_sheet)

    # 3번째 요구사항 excel의 select_write_sheet(입력받은 sheet) sheet에 요구사항을 실행한다
    make_third_excel(
        write_excel, load_excel_file, select_write_sheet)

    # 3 번째 요구에맞는 엑셀만들기
    write_excel.save(output_xcel_src)
