from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
# 모듈을 현재 디렉토리에서 import합니다.
from src.make_base_sheet.second_make_base import make_second_base
from src.make_excel.second_make_excel import make_second_excel
import String

# 두번째 요구의 sheet 이름
second_1차상담_직탐미설 = String.second_1차상담_직탐미설
second_1차상담_신입생세미나 = String.second_1차상담_신입생세미나
second_1차상담_자기주도 = String.second_1차상담_자기주도
second_1차상담_진로집단상담 = String.second_1차상담_진로집단상담
second_1차상담_단기직무체험 = String.second_1차상담_단기직무체험

second_2차상담_직탐미설 = String.second_2차상담_직탐미설
second_2차상담_포트워크숍 = String.second_2차상담_포트워크숍
second_2차상담_자기주도 = String.second_2차상담_자기주도
second_2차상담_진로집단상담 = String.second_2차상담_진로집단상담
second_2차상담_단기직무체험 = String.second_2차상담_단기직무체험

second_3차상담_포트폴리오컨설팅 = String.second_3차상담_포트폴리오컨설팅

base_sheet_list = [
    second_1차상담_직탐미설,
    second_1차상담_신입생세미나,
    second_1차상담_자기주도,
    second_1차상담_진로집단상담,
    second_1차상담_단기직무체험,

    second_2차상담_직탐미설,
    second_2차상담_포트워크숍,
    second_2차상담_자기주도,
    second_2차상담_진로집단상담,
    second_2차상담_단기직무체험,

    second_3차상담_포트폴리오컨설팅
]


def make_second_request_excel():
    # second file 이름
    inputFile = String.NAME_SECOND_INPUT_FILE + ".xlsx"

    outputFile = String.NAME_SECOND_OUT_FILE + ".xlsx"

    input_xcel_src = "./input_excel_file/" + inputFile
    output_xcel_src = "./output_excel_file/" + outputFile

    # 엑셀파일 쓰기
    write_excel = Workbook()

    # read xcel file
    # data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
    load_excel_file = load_workbook(
        input_xcel_src, data_only=True)

    # 기본 시트 제거
    default_sheet = write_excel.active
    write_excel.remove(default_sheet)

    for select_write_sheet in base_sheet_list:
        # 2 번째 sheet 만들기
        write_excel.create_sheet(select_write_sheet)

        # 2번째 sheet 열기
        sheet_second_write_excel = write_excel[select_write_sheet]

        # 2번째요구의  시트의 베이스 엑셀을 만든다.
        make_second_base(
            sheet_second_write_excel, String.num_width_colum)

    # 2번째 요구사항 excel의 2번째 sheet에 요구사항을 실행한다
    make_second_excel(
        write_excel, load_excel_file)

    # 2 번째 요구에맞는 엑셀만들기
    write_excel.save(output_xcel_src)
